#!/usr/bin/python
# -*- coding: UTF-8 -*-

import csv
import codecs
import json
import openpyxl
import xlwt,xlrd
from xlutils import copy as xl_copy


# x = open('student.txt')
# # x.readline() 读取1行
# # x.read 读取所有文本
# stringtxt = x.read()  #因为student.txt存在中文报错。
# print(stringtxt)

#1 解决读取中问题
strf = ''
f = codecs.open('student.txt','r','utf-8')
s = f.readlines()
f.close()
for line in s:
    strf = strf + line
print(strf)

#2 json转字典
jsonf = json.loads(strf)
print(jsonf["1"])
print(jsonf["1"][0])

#3 存xlsx


# with codecs.open('student.txt','r','utf-8') as tab_file:
#     tab_reader = csv.reader(tab_file, delimiter='\t')
#     xls_readable_book = xlrd.open_workbook('student.xlsx')
#     xls_writeable_book = xl_copy.copy(xls_readable_book)
#     xls_writeable_sheet = xls_writeable_book.get_sheet(0)
#     for row_index, row in jsonf:
#         xls_writeable_sheet.write(row_index, 0, row[0])
#         xls_writeable_sheet.write(row_index, 1, row[1])
#     xls_writeable_book.save('stack37.xls')


#新建excel
def creatwb(wbname):
    wb = openpyxl.Workbook()
    wb.save(filename=wbname)
    print("新建Excel："+wbname+"成功")

# 写入excel文件中rows==startrow注意向下写要自增, date 数据，date是list数据类型， fields 表头
def savetoexcel(rows,data,fields,sheetname,wbname):
    print("写入excel：")
    wb = openpyxl.load_workbook(filename=wbname)

    sheet = wb.active
    sheet.title = sheetname

    field = 1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1 = 1
    for row1 in range(1,len(data)+1):  # 写入数据
        # for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=rows,column=row1,value=str(data[row1-1]))

    wb.save(filename=wbname)
    print("保存成功")

for index in jsonf:
    savetoexcel(int(index)+1,jsonf[index],["姓名","数学","语文","计算机"],"sheet4","student.xlsx")
